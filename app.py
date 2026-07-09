from flask import Flask, json, render_template, request, redirect, url_for, session, send_file
import pandas as pd
from io import BytesIO
from reportlab.pdfgen import canvas
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

app = Flask(__name__)
app.secret_key = "sena2026"

# ==========================
# Usuarios de prueba (Simulados con Roles)
# ==========================
usuarios = {
    "admin@sena.edu.co": {
        "password": "123",
        "rol": "admin"
    },
    "aprendiz@sena.edu.co": {
        "password": "123",
        "rol": "aprendiz",
        "ficha": "3407184"  # <-- Le asignamos una ficha fija para que tenga lógica de roles
    }
}

# ==========================
# Reportes de prueba (Simulados con predeterminados)
# ==========================
datos_evaluaciones = [
    {
        "Instructor": "Carlos Pérez",
        "Programa": "ADSO",
        "Ficha": "2876541",
        "Puntaje": 4.8,
        "Comentario": "Excelente dominio del tema"
    },
    {
        "Instructor": "Ana Torres",
        "Programa": "Contabilidad",
        "Ficha": "2987654",
        "Puntaje": 4.5,
        "Comentario": "Muy buena metodología"
    },
    {
        "Instructor": "Luis Gómez",
        "Programa": "Diseño",
        "Ficha": "3054789",
        "Puntaje": 4.2,
        "Comentario": "Explica claramente"
    }
]

# ==========================
# Home
# ==========================
@app.route('/')
def home():
    return render_template('home.html')


# ==========================
# Login
# ==========================
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        correo = request.form['correo']
        password = request.form['password']

        if correo in usuarios and usuarios[correo]["password"] == password:
            session["usuario"] = correo
            session["rol"] = usuarios[correo]["rol"]
            
            # Si es aprendiz, guardamos su número de ficha en la sesión
            if usuarios[correo]["rol"] == "aprendiz":
                session["ficha"] = usuarios[correo]["ficha"]
                
            return redirect(url_for('home'))
        else:
            return render_template("login.html", error="Correo o contraseña incorrectos.")
    return render_template("login.html")

# ==========================
# Logout
# ==========================
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ==========================
# Instructores
# ==========================
@app.route('/instructores')
def instructores():
    if not session.get("usuario"):
        return redirect(url_for("login"))
    return render_template('instructores.html')


@app.route('/crear_instructor')
def crear_instructor():
    if session.get("rol") != "admin":
        return redirect(url_for("home"))
    return render_template('crear_instructor.html')


@app.route('/actualizar_instructor')
def actualizar_instructor():
    if session.get("rol") != "admin":
        return redirect(url_for("home"))
    return render_template('actualizar_instructor.html')


@app.route('/eliminar_instructor')
def eliminar_instructor():
    if session.get("rol") != "admin":
        return redirect(url_for("home"))
    return render_template('eliminar_instructor.html')


# ==========================
# Contacto
# ==========================
@app.route('/contacto')
def contacto():
    return render_template('contacto.html')


# ==========================
# Evaluaciones
# ==========================
@app.route('/evaluaciones')
def evaluaciones():
    if not session.get("usuario"):
        return redirect(url_for("login"))
    return render_template("evaluaciones.html")


# NUEVA RUTA: Modificar preguntas (Solo Admin)
@app.route('/evaluaciones/editar')
def editar_preguntas():
    if session.get("rol") != "admin":
        return redirect(url_for("home"))
    # Aquí debes asegurarte de tener un archivo llamado 'editar_preguntas.html' dentro de templates/
    return render_template("editar_preguntas.html")


# ==========================
# Fichas
# ==========================
@app.route('/fichas')
def fichas():
    if session.get("rol") != "admin":
        return redirect(url_for("home"))
    return render_template('fichas.html')


# ==========================
# Reportes
# ==========================
@app.route('/reportes')
def reportes():
    if session.get("rol") != "admin":
        return redirect(url_for("home"))
    return render_template('reportes.html')

@app.route("/descargar_excel")
def descargar_excel():

    df = pd.DataFrame(datos_evaluaciones)
    archivo = BytesIO()

    with pd.ExcelWriter(archivo, engine="openpyxl") as writer:

        df.to_excel(writer, startrow=2, index=False)

        hoja = writer.sheets["Sheet1"]

        hoja["A1"] = "REPORTE DE EVALUACIONES DE INSTRUCTORES"
        hoja["A1"].font = Font(size=16, bold=True)
        hoja["A1"].alignment = Alignment(horizontal="center")

        hoja.merge_cells("A1:E1")

        hoja["A2"] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

        verde = PatternFill(fill_type="solid", fgColor="00843D")

        for celda in hoja[3]:
            celda.fill = verde
            celda.font = Font(color="FFFFFF", bold=True)
            celda.alignment = Alignment(horizontal="center")

    archivo.seek(0)

    return send_file(
        archivo,
        as_attachment=True,
        download_name="Reporte_Evaluaciones.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/descargar_pdf")
def descargar_pdf():

    archivo = BytesIO()
    pdf = canvas.Canvas(archivo)

    # Título del documento
    pdf.setTitle("Reporte de Evaluaciones")

    # Encabezado
    pdf.setFont("Helvetica-Bold", 18)
    pdf.drawString(120, 810, "REPORTE DE EVALUACIONES")

    pdf.setFont("Helvetica", 12)
    pdf.drawString(40, 790, "Sistema de Evaluación de Instructores")
    pdf.drawString(40, 775, "Servicio Nacional de Aprendizaje - SENA")
    pdf.drawString(
        40,
        760,
        f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )

    # Línea separadora
    pdf.line(40, 750, 560, 750)

    # Encabezados de la tabla
    pdf.setFont("Helvetica-Bold", 10)

    pdf.drawString(40, 730, "Instructor")
    pdf.drawString(170, 730, "Programa")
    pdf.drawString(290, 730, "Ficha")
    pdf.drawString(360, 730, "Puntaje")
    pdf.drawString(430, 730, "Comentario")

    pdf.line(40, 720, 560, 720)

    # Posición inicial
    y = 700

    pdf.setFont("Helvetica", 10)

    # Datos
    for e in datos_evaluaciones:

        pdf.drawString(40, y, str(e["Instructor"]))
        pdf.drawString(170, y, str(e["Programa"]))
        pdf.drawString(290, y, str(e["Ficha"]))
        pdf.drawString(360, y, str(e["Puntaje"]))
        pdf.drawString(430, y, str(e["Comentario"]))

        y -= 25

        # Nueva página si se acaba el espacio
        if y < 70:
            pdf.showPage()

            pdf.setFont("Helvetica-Bold", 10)

            pdf.drawString(40, 800, "Instructor")
            pdf.drawString(170, 800, "Programa")
            pdf.drawString(290, 800, "Ficha")
            pdf.drawString(360, 800, "Puntaje")
            pdf.drawString(430, 800, "Comentario")

            pdf.line(40, 790, 560, 790)

            pdf.setFont("Helvetica", 10)

            y = 770

    # Pie de página
    pdf.setFont("Helvetica-Oblique", 9)
    pdf.drawString(
        40,
        40,
        "Documento generado automáticamente por el Sistema de Evaluación de Instructores."
    )

    # Guardar el PDF
    pdf.save()

    # Regresar al inicio del archivo
    archivo.seek(0)

    return send_file(
        archivo,
        as_attachment=True,
        download_name="Reporte_Evaluaciones.pdf",
        mimetype="application/pdf"
    )

# ==========================
# Ejecutar aplicación
# ==========================
if __name__ == '__main__':
    app.run(debug=True)
